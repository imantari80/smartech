"""
export_estadisticas.py
=======================
Exporta a Excel las estadísticas de uso del chatbot SmartTech
(cantidad de consultas / conversaciones), agrupadas por día, semana o mes.

Requiere:
- La tabla `chat_log` creada (ver migracion_estadisticas.sql)
- Que routes.py esté registrando cada mensaje con chat_logger.registrar_evento()
- openpyxl (ya usado en el proyecto para otras exportaciones)

Uso típico (desde una ruta Flask, ver ejemplo al final del archivo):

    from export_estadisticas import generar_reporte_excel
    buffer = generar_reporte_excel(fecha_inicio, fecha_fin, agrupacion='dia')
    return send_file(buffer, ..., as_attachment=True, download_name='reporte.xlsx')
"""

import io
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from services import obtener_conexion

# ------------------------------------------------------------------
# Agrupaciones soportadas: 'dia', 'semana', 'mes'
# ------------------------------------------------------------------
_TRUNC_POSTGRES = {'dia': 'day', 'semana': 'week', 'mes': 'month'}
_ETIQUETAS = {'dia': 'Día', 'semana': 'Semana', 'mes': 'Mes'}
_FORMATOS_FECHA = {'dia': 'dd/mm/yyyy', 'semana': 'dd/mm/yyyy', 'mes': 'mmm yyyy'}


def _validar_agrupacion(agrupacion):
    a = (agrupacion or 'dia').strip().lower()
    if a not in _TRUNC_POSTGRES:
        raise ValueError("El parámetro 'agrupacion' debe ser 'dia', 'semana' o 'mes'.")
    return a


def _validar_rango(fecha_inicio, fecha_fin):
    if fecha_inicio is None or fecha_fin is None:
        raise ValueError("Debes indicar fecha_inicio y fecha_fin.")
    if fecha_fin <= fecha_inicio:
        raise ValueError("fecha_fin debe ser posterior a fecha_inicio.")


# ------------------------------------------------------------------
# Consultas a la base de datos
# ------------------------------------------------------------------

def _totales_generales(cursor, fecha_inicio, fecha_fin):
    """Total de consultas (mensajes) y de conversaciones (sesiones) únicas
    en TODO el rango. Se calcula aparte porque sumar 'distinct' por período
    sobreestimaría el total si una misma sesión escribe en más de un período."""
    cursor.execute(
        """
        SELECT COUNT(*) AS total_consultas,
               COUNT(DISTINCT sesion_id) AS conversaciones_unicas
        FROM chat_log
        WHERE fecha_hora >= %s AND fecha_hora < %s;
        """,
        (fecha_inicio, fecha_fin)
    )
    return cursor.fetchone()  # (total_consultas, conversaciones_unicas)


def _totales_por_periodo(cursor, fecha_inicio, fecha_fin, agrupacion):
    trunc = _TRUNC_POSTGRES[agrupacion]
    cursor.execute(
        f"""
        SELECT date_trunc(%s, fecha_hora) AS periodo,
               COUNT(*) AS total_consultas,
               COUNT(DISTINCT sesion_id) AS conversaciones_unicas
        FROM chat_log
        WHERE fecha_hora >= %s AND fecha_hora < %s
        GROUP BY periodo
        ORDER BY periodo;
        """,
        (trunc, fecha_inicio, fecha_fin)
    )
    return cursor.fetchall()  # [(periodo, total_consultas, conversaciones_unicas), ...]


def _totales_por_perfil(cursor, fecha_inicio, fecha_fin):
    cursor.execute(
        """
        SELECT COALESCE(perfil, 'Sin definir') AS perfil, COUNT(*) AS total
        FROM chat_log
        WHERE fecha_hora >= %s AND fecha_hora < %s
        GROUP BY perfil
        ORDER BY total DESC;
        """,
        (fecha_inicio, fecha_fin)
    )
    return cursor.fetchall()


def _totales_por_formato(cursor, fecha_inicio, fecha_fin):
    cursor.execute(
        """
        SELECT COALESCE(formato, 'Sin definir') AS formato, COUNT(*) AS total
        FROM chat_log
        WHERE fecha_hora >= %s AND fecha_hora < %s
        GROUP BY formato
        ORDER BY total DESC;
        """,
        (fecha_inicio, fecha_fin)
    )
    return cursor.fetchall()


def _detalle_crudo(cursor, fecha_inicio, fecha_fin, limite=20000):
    cursor.execute(
        """
        SELECT fecha_hora, sesion_id, paso, evento, perfil, formato, presupuesto, mensaje_usuario
        FROM chat_log
        WHERE fecha_hora >= %s AND fecha_hora < %s
        ORDER BY fecha_hora
        LIMIT %s;
        """,
        (fecha_inicio, fecha_fin, limite)
    )
    return cursor.fetchall()


# ------------------------------------------------------------------
# Estilos reutilizables para el Excel
# ------------------------------------------------------------------
FONT_NAME = "Arial"
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, size=10, bold=True)
KPI_FONT = Font(name=FONT_NAME, size=20, bold=True, color="1F4E78")
KPI_LABEL_FONT = Font(name=FONT_NAME, size=10, color="595959")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
KPI_FILL = PatternFill("solid", fgColor="D9E1F2")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(vertical="center", horizontal="center")


def _titulo_hoja(ws, texto, num_columnas):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columnas)
    c = ws.cell(row=1, column=1, value=texto)
    c.font = TITLE_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28


def _kpi_card(ws, col, valor, etiqueta):
    letra = get_column_letter(col)
    ws.merge_cells(f"{letra}3:{letra}4")
    c = ws[f"{letra}3"]
    c.value = valor
    c.font = KPI_FONT
    c.fill = KPI_FILL
    c.alignment = CENTER
    ws.merge_cells(f"{letra}5:{letra}5")
    c2 = ws[f"{letra}5"]
    c2.value = etiqueta
    c2.font = KPI_LABEL_FONT
    c2.fill = KPI_FILL
    c2.alignment = CENTER
    ws.row_dimensions[3].height = 34


def resumen_dashboard():
    """
    Números rápidos para las tarjetas del dashboard (hoy / esta semana /
    este mes / conversaciones únicas del mes) + una serie diaria de los
    últimos 14 días para el gráfico. Pensado para ser liviano: se llama
    cada vez que el administrador abre el panel.
    """
    hoy = datetime.now().date()
    inicio_hoy = datetime.combine(hoy, datetime.min.time())
    fin_hoy = inicio_hoy + timedelta(days=1)
    inicio_semana = inicio_hoy - timedelta(days=hoy.weekday())  # lunes de esta semana
    inicio_mes = inicio_hoy.replace(day=1)
    inicio_serie = inicio_hoy - timedelta(days=13)

    conn = obtener_conexion()
    cursor = conn.cursor()
    try:
        total_hoy, _ = _totales_generales(cursor, inicio_hoy, fin_hoy)
        total_semana, _ = _totales_generales(cursor, inicio_semana, fin_hoy)
        total_mes, conversaciones_mes = _totales_generales(cursor, inicio_mes, fin_hoy)
        serie = _totales_por_periodo(cursor, inicio_serie, fin_hoy, 'dia')
    finally:
        cursor.close()
        conn.close()

    return {
        "hoy": total_hoy,
        "semana": total_semana,
        "mes": total_mes,
        "conversaciones_mes": conversaciones_mes,
        "serie": [
            {"fecha": periodo.strftime('%Y-%m-%d'), "total": total}
            for periodo, total, _conversaciones in serie
        ],
    }


# ------------------------------------------------------------------
# Construcción del workbook
# ------------------------------------------------------------------

def generar_reporte_excel(fecha_inicio, fecha_fin, agrupacion='dia'):
    """
    Genera el reporte de estadísticas del chatbot en memoria (BytesIO).

    Parámetros
    ----------
    fecha_inicio : datetime  (incluido)
    fecha_fin    : datetime  (excluido, es decir el rango es [inicio, fin) )
    agrupacion   : 'dia' | 'semana' | 'mes'

    Retorna
    -------
    io.BytesIO listo para enviarse con Flask send_file().
    """
    agrupacion = _validar_agrupacion(agrupacion)
    _validar_rango(fecha_inicio, fecha_fin)

    conn = obtener_conexion()
    cursor = conn.cursor()
    try:
        total_consultas, total_conversaciones = _totales_generales(cursor, fecha_inicio, fecha_fin)
        por_periodo = _totales_por_periodo(cursor, fecha_inicio, fecha_fin, agrupacion)
        por_perfil = _totales_por_perfil(cursor, fecha_inicio, fecha_fin)
        por_formato = _totales_por_formato(cursor, fecha_inicio, fecha_fin)
        detalle = _detalle_crudo(cursor, fecha_inicio, fecha_fin)
    finally:
        cursor.close()
        conn.close()

    dias_rango = max((fecha_fin - fecha_inicio).days, 1)
    promedio_diario = round(total_consultas / dias_rango, 1)

    wb = openpyxl.Workbook()

    # ================= HOJA 1: RESUMEN =================
    ws = wb.active
    ws.title = "Resumen"
    _titulo_hoja(
        ws,
        f"SMARTECH — Estadísticas del Chatbot ({fecha_inicio:%d/%m/%Y} al "
        f"{(fecha_fin - timedelta(days=1)):%d/%m/%Y}) — Agrupado por {_ETIQUETAS[agrupacion]}",
        6
    )

    _kpi_card(ws, 1, total_consultas, "Total de consultas (mensajes)")
    _kpi_card(ws, 3, total_conversaciones, "Conversaciones únicas")
    _kpi_card(ws, 5, promedio_diario, "Promedio de consultas / día")

    fila = 7
    ws.cell(row=fila, column=1, value=_ETIQUETAS[agrupacion]).font = HEADER_FONT
    ws.cell(row=fila, column=2, value="Total de consultas").font = HEADER_FONT
    ws.cell(row=fila, column=3, value="Conversaciones únicas").font = HEADER_FONT
    for col in (1, 2, 3):
        cell = ws.cell(row=fila, column=col)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    fila += 1
    primera_fila_datos = fila

    for i, (periodo, total, conversaciones) in enumerate(por_periodo):
        ws.cell(row=fila, column=1, value=periodo).number_format = _FORMATOS_FECHA[agrupacion]
        ws.cell(row=fila, column=2, value=total)
        ws.cell(row=fila, column=3, value=conversaciones)
        for col in (1, 2, 3):
            cell = ws.cell(row=fila, column=col)
            cell.font = NORMAL_FONT
            cell.border = BORDER
            cell.alignment = CENTER
            if i % 2 == 0:
                cell.fill = ALT_FILL
        fila += 1

    ultima_fila_datos = fila - 1
    if ultima_fila_datos >= primera_fila_datos:
        # Totales de control, calculados con fórmula (deben coincidir con el KPI)
        ws.cell(row=fila, column=1, value="TOTAL").font = BOLD_FONT
        ws.cell(row=fila, column=2,
                value=f"=SUM(B{primera_fila_datos}:B{ultima_fila_datos})").font = BOLD_FONT
        for col in (1, 2, 3):
            ws.cell(row=fila, column=col).border = BORDER

        # Gráfico de barras con la evolución de consultas por período
        chart = BarChart()
        chart.title = f"Consultas por {_ETIQUETAS[agrupacion].lower()}"
        chart.y_axis.title = "Consultas"
        chart.x_axis.title = _ETIQUETAS[agrupacion]
        datos = Reference(ws, min_col=2, min_row=7, max_row=ultima_fila_datos)
        categorias = Reference(ws, min_col=1, min_row=primera_fila_datos, max_row=ultima_fila_datos)
        chart.add_data(datos, titles_from_data=True)
        chart.set_categories(categorias)
        chart.width = 18
        chart.height = 9
        ws.add_chart(chart, f"E{primera_fila_datos}")

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 22
    ws.freeze_panes = f"A{primera_fila_datos}"

    # ================= HOJA 2: POR PERFIL =================
    ws2 = wb.create_sheet("Por Perfil")
    _titulo_hoja(ws2, "Consultas por perfil de usuario", 2)
    ws2.cell(row=3, column=1, value="Perfil").font = HEADER_FONT
    ws2.cell(row=3, column=2, value="Total de consultas").font = HEADER_FONT
    for col in (1, 2):
        c = ws2.cell(row=3, column=col)
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    r = 4
    for i, (perfil, total) in enumerate(por_perfil):
        ws2.cell(row=r, column=1, value=perfil).font = NORMAL_FONT
        ws2.cell(row=r, column=2, value=total).font = NORMAL_FONT
        for col in (1, 2):
            cell = ws2.cell(row=r, column=col)
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL
        r += 1
    ws2.column_dimensions['A'].width = 34
    ws2.column_dimensions['B'].width = 20

    # ================= HOJA 3: POR FORMATO =================
    ws3 = wb.create_sheet("Por Formato")
    _titulo_hoja(ws3, "Consultas por formato de equipo", 2)
    ws3.cell(row=3, column=1, value="Formato").font = HEADER_FONT
    ws3.cell(row=3, column=2, value="Total de consultas").font = HEADER_FONT
    for col in (1, 2):
        c = ws3.cell(row=3, column=col)
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    r = 4
    for i, (formato, total) in enumerate(por_formato):
        ws3.cell(row=r, column=1, value=formato).font = NORMAL_FONT
        ws3.cell(row=r, column=2, value=total).font = NORMAL_FONT
        for col in (1, 2):
            cell = ws3.cell(row=r, column=col)
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL
        r += 1
    ws3.column_dimensions['A'].width = 28
    ws3.column_dimensions['B'].width = 20

    # ================= HOJA 4: DETALLE (log crudo) =================
    ws4 = wb.create_sheet("Detalle")
    _titulo_hoja(ws4, "Detalle de cada consulta (máx. 20,000 filas)", 8)
    headers = ["Fecha y hora", "Sesión", "Paso", "Evento", "Perfil", "Formato", "Presupuesto", "Mensaje"]
    for c, h in enumerate(headers, start=1):
        cell = ws4.cell(row=3, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    r = 4
    for i, row in enumerate(detalle):
        fecha_hora, sesion_id, paso, evento, perfil, formato, presupuesto, mensaje = row
        valores = [fecha_hora, sesion_id, paso, evento, perfil, formato, presupuesto, mensaje]
        for c, v in enumerate(valores, start=1):
            cell = ws4.cell(row=r, column=c, value=v)
            cell.font = NORMAL_FONT
            cell.border = BORDER
            cell.alignment = WRAP
            if c == 1:
                cell.number_format = "dd/mm/yyyy hh:mm"
            if i % 2 == 0:
                cell.fill = ALT_FILL
        r += 1
    anchos = {"A": 18, "B": 26, "C": 8, "D": 18, "E": 22, "F": 18, "G": 14, "H": 50}
    for col, w in anchos.items():
        ws4.column_dimensions[col].width = w
    ws4.freeze_panes = "A4"
    if r > 4:
        ws4.auto_filter.ref = f"A3:H{r - 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ------------------------------------------------------------------
# Ejemplo de integración como ruta Flask (agregar en routes.py)
# ------------------------------------------------------------------
"""
from export_estadisticas import generar_reporte_excel
from datetime import datetime, timedelta

@chat_bp.route('/api/estadisticas/exportar')
def exportar_estadisticas():
    agrupacion = request.args.get('agrupacion', 'dia')      # 'dia' | 'semana' | 'mes'
    hasta_str  = request.args.get('hasta')                  # 'YYYY-MM-DD', opcional
    desde_str  = request.args.get('desde')                  # 'YYYY-MM-DD', opcional

    hoy = datetime.now().date()
    fecha_fin = (datetime.strptime(hasta_str, '%Y-%m-%d').date() + timedelta(days=1)
                 if hasta_str else hoy + timedelta(days=1))
    fecha_inicio = (datetime.strptime(desde_str, '%Y-%m-%d').date()
                    if desde_str else fecha_fin - timedelta(days=30))

    try:
        buffer = generar_reporte_excel(
            datetime.combine(fecha_inicio, datetime.min.time()),
            datetime.combine(fecha_fin, datetime.min.time()),
            agrupacion=agrupacion,
        )
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    nombre = f"Estadisticas_Chatbot_{fecha_inicio}_a_{fecha_fin - timedelta(days=1)}.xlsx"
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre,
    )
"""
